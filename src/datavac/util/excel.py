from __future__ import annotations
from typing import TYPE_CHECKING, Iterator, Optional
from contextlib import contextmanager
from pathlib import Path

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook


@contextmanager
def cache_wbs():
    """ Context-manager to cache the results of openpyxl's load_workbook.

    Any use of openpyxl's load_workbook within this context will be cached,
    and the returned workbooks' close() function will also be monkey-patched
    to do nothing.  When the context is exited, any of the workbooks opened
    within the context will be closed.
    """
    import openpyxl
    wb_cache={}
    true_load_workbook=openpyxl.load_workbook
    def cached_load_workbook(file,**kwargs):
        if not (type(file) is str or isinstance(file,Path)):
            return true_load_workbook(file,**kwargs)
        if str(Path(file)) not in wb_cache:
            wb_cache[str(Path(file))]=true_load_workbook(file,read_only=True)
            wb_cache[str(Path(file))]._true_close=wb_cache[str(Path(file))].close
            wb_cache[str(Path(file))].close=lambda *args,**kwargs: None
        return wb_cache[str(Path(file))]
    openpyxl.load_workbook=cached_load_workbook
    try: yield
    finally:
        openpyxl.load_workbook=true_load_workbook
        for wb in wb_cache.values():
            wb._true_close()


@contextmanager
def open_workbook_for_read(file: Path|str, preopened_wb: Optional[Workbook] = None) -> Iterator[Workbook]:
    """ Context manager like python's open() but using openpyxl.

        Args:
            file (path-like): path to open
            preopened_wb (Workbook): if supplied, preopened_wb be yielded instead of opening file,
                and this context will not be responsible for closing it.

        Yields:
            the Workbook (or preopened_wb)
    """
    import openpyxl
    if preopened_wb is not None:
        yield preopened_wb
    else:
        wb=openpyxl.load_workbook(file,read_only=True)
        try: yield wb
        finally: wb.close()